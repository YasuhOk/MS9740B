import openpyxl as excel
from datetime import datetime
import time
from datetime import datetime as dt
from datetime import timedelta
import pandas as pd
import pickle
import os
import PySimpleGUI as sg
import configparser
from bs4 import BeautifulSoup
from datetime import date as days
import webbrowser
import shutil
from configparser import ConfigParser
import json
import subprocess
import pyautogui as pgui
import pyperclip as clip
import datetime
import ctypes

from pyscreeze import showRegionOnScreen
# import R3AutoExeMain_002 as R3002
# import R3AutoExeMain_003 as R3003

# ------------------------------------------------------
# 関数名     get_first_date
# 用途       月初を取得
# 引数       dt = 日付
# ------------------------------------------------------
def get_first_date(dt):
    return dt.replace(day=1)
# ------------------------------------------------------
# 関数名     inijson
# 用途       iniファイルをjsonに変換
# 引数       json_path = jsonファイルパス_出力
#            ini_path = iniファイルパス_読込
# ------------------------------------------------------
def inijson(json_path, ini_path):
    d = {}
    cfg = ConfigParser()
    cfg.read(ini_path,encoding="utf-8")
    for s in cfg.sections():
        # print(s)
        d[s] = dict(cfg.items(s))
    with open(json_path, 'w', encoding="utf-8") as f:
        json.dump(d, f, ensure_ascii=False)
# inijson("X:\Python\MS9740B進捗管理アプリケーション\Config.json","X:\Python\MS9740B進捗管理アプリケーション\Config.ini")
# ------------------------------------------------------
# 関数名     ExcelSheetHitColumn
# 用途       シートのデータの行を指定し検索条件にマッチした列を返す
# 引数       sheet = 読込後のシートを指定する
#            StartRow = 取得開始縦列 int
#            FirstMachString = 検索第一条件
# 戻り値     Column
# ------------------------------------------------------
def ExcelSheetHitColumn(sheet,intStartRow,FirstMachString):
    # 範囲データを指定行のみ取得
    for row in sheet.iter_rows(min_row=intStartRow,max_row=intStartRow):
        # 指定行の列、セルを取得
        for index , cell in enumerate(row,start=1):
            # 検索条件に有ったColumnを返す
            if cell.value == FirstMachString:
                return cell.column
# ------------------------------------------------------
# 関数名     CreateDic
# 用途       見出しセルからメインとなるKeyを指定し多次元辞書型にデータを格納する
# 引数       sheet = シートを指定
#            MainKey = 主となる見出し
#            SearchKey[] = 検索見出しをリストで指定
#           HeadStartRow = 見出し開始行指定
#           DataGetStartRow = データ取得開始行指定
# 戻り値     多次元辞書型配列
# ------------------------------------------------------
def CreateDic(sheet,MainKey,listSearchKey,HeadStartRow,DataGetStartRow):
    # 辞書型、配列を初期化
    valueDat = {} 
    HitCol = []
    # メインキーのColumnを返す
    MainKeyCol = ExcelSheetHitColumn(sheet,HeadStartRow,MainKey)
    # 指定データのColumnを返す_配列に格納
    for index in range(len(listSearchKey)):
        HitCol.append(ExcelSheetHitColumn(sheet,HeadStartRow,listSearchKey[index]))
    # シート上にある全データを1行ずつループ
    for row in sheet.iter_rows(min_row=DataGetStartRow):
        # 1行毎にセル1列ずつ値を取得し配列に格納
        values = [cell.value for cell in row]
        # メインキー列が空じゃない場合
        if values[MainKeyCol-1] is not None:
            # メインキーを取り出す 横列指定
            strMainKey = str(values[MainKeyCol-1])
            # 初出ならリスト初期化
            if strMainKey not in valueDat: valueDat[strMainKey] = {}
            # 多次元辞書型に格納する
            for index , item in enumerate(HitCol):
                valueDat[strMainKey].update({listSearchKey[index]:values[item-1]})
    return valueDat
# ------------------------------------------------------
# 関数名     CreateDicDay
# 用途       見出しセルから日付データを辞書型にデータを格納する
# 引数      sheet = シートを指定
#           StartRowSet = データ取得開始行指定
# 戻り値     辞書型配列
# ------------------------------------------------------
def CreateDicDay(sheet,StartRowSet):
    # 辞書型、配列を初期化
    dicColDay = {} 
    # 範囲データを指定行のみ取得
    for row in sheet.iter_rows(min_row=StartRowSet,max_row=StartRowSet):
        # 指定行の列、セルを取得
        for index , cell in enumerate(row,start=1):
            # セルが日付の場合辞書型にデータを格納
            if isinstance(cell.value, datetime):
                # 初出の場合辞書型に代入する
                if index not in dicColDay: 
                    dicColDay[index] = {}
                    dicColDay[index+1] = {}
                # 日付をフォーマット形式指定し代入する＿dicColDay=Column:{日付:yyyy/mm/dd}
                dicColDay[index].update({"日付":"{0:%Y/%m/%d}".format(cell.value)})
                dicColDay[index+1].update({"日付":"{0:%Y/%m/%d}".format(cell.value)})
    return dicColDay
# ------------------------------------------------------
# 関数名     UPDateDicData
# 用途       辞書型配列をアップデート
# 引数      DicvalueDat = 元データ
#           UpdateKey = 更新キー
#           strName = 更新名
#           Col = 日付列
#           strPatternType = 塗りつぶしタイプ
#           intfgColor = セル色判定
#           intsgColor = セル色判定
# 戻り値     辞書型配列
# ------------------------------------------------------
def UPDateDicData(dicColDay,DicvalueDat,UpdateKey,strName,Col,strPatternType,intfgColor,intsgColor):
    SasizuDay = dicColDay[Col]["日付"]
    DicvalueDat[UpdateKey].update({strName + "_日付":SasizuDay})
    if strPatternType == "solid" and intfgColor == 0 and intsgColor == 64:
        DicvalueDat[UpdateKey].update({strName + "_進捗":"完"})
    else:
        DicvalueDat[UpdateKey].update({strName + "_進捗":"未"})
    return DicvalueDat
# ------------------------------------------------------
# 関数名     isint
# 用途       文字列が数値を表し、int／float関数による変換が可能かどうかを判定
# 引数      strDat = 元データ
# 戻り値     True/False
# ------------------------------------------------------
def isint(strDat):
    try:
        int(strDat) 
    except ValueError:
        return False
    else:
        return True

if __name__ == "__main__":
    # デザインテーマの設定
    sg.theme("BrownBlue")  
    # フォント、サイズ指定
    font = ("meiryo",10)
    # Config.iniファイルからパス設定を取得
    config = configparser.ConfigParser()
    ConfigIniPath = os.path.join(os.path.dirname(os.path.abspath(__file__)) ,"Config.ini")
    config.read(ConfigIniPath, encoding="utf-8")
    SGExcelPath = config.get("Path","SGExcelPath")
    MSExcelPath = config.get("Path","MSExcelPath")
    HtmlOutPath = config.get("Path","htmloutpath")
    ProvisionalNumber = config.get("DefaultSet","ProvisionalNumber")
    KougakubuPath = config.get("Path","kougakubupath")
    KanryouPath = config.get("Path","kanryoupath")
    ActualTimepath = config.get("Path","actualtimepath")

    SetButtonColor="#3d3d35"
    # ウィンドウに配置するコンポーネント--------------------------------------------------------------------------------------------------------------------
    frame1 =[[sg.Radio(text="【全仮番取得】指定仮番号全て取得", group_id="A", default=False,text_color="#d0e4f3",font=font)],
            [sg.Radio(text="【仕掛品取得】既出荷品は取得しない", group_id="A", default=True,text_color="#d0e4f3",font=font)]]
    col1 = [[sg.Text("取得したい開始仮番号を入力してください",text_color="#d0e4f3",font=font)],
            [sg.InputText(font=font,default_text=ProvisionalNumber,size=(7,10),background_color="#d0e4f3")]]
    layout =[[sg.Text("----------------------------------------------------------------------------------------",text_color="#d0e4f3",font=font)],
            [sg.FileBrowse("計画進捗表",font=font) ,sg.InputText(font=font,default_text=SGExcelPath,background_color="#02497c",text_color="#ffffff",size=(55,1))],
            [sg.FileBrowse("製造進捗表",font=font) ,sg.InputText(font=font,default_text=MSExcelPath,background_color="#02497c",text_color="#ffffff",size=(55,1))],
            [sg.FolderBrowse("HTML出力 ",font=font) ,sg.InputText(font=font,default_text=HtmlOutPath,background_color="#02497c",text_color="#ffffff",size=(55,1))],
            [sg.Frame("取得モード",frame1),sg.Frame("仮番号指定",col1)],
            [sg.Button("更新",button_color=SetButtonColor), sg.Button("出力",button_color=SetButtonColor)],
            [sg.Text("----------------------------------------------------------------------------------------",text_color="#d0e4f3",font=font)],
            [sg.FileBrowse("光学部管理",font=font) ,sg.InputText(font=font,default_text=KougakubuPath,background_color="#02497c",text_color="#ffffff",size=(55,1))],
            [sg.FileBrowse("完了管理表",font=font) ,sg.InputText(font=font,default_text=KanryouPath,background_color="#02497c",text_color="#ffffff",size=(55,1))],
            [sg.Button("光学部開く",button_color=SetButtonColor), sg.Button("完了表開く",button_color=SetButtonColor), sg.Button("R/3(002)実行",button_color=SetButtonColor), sg.Button("R/3(003)実行",button_color=SetButtonColor)],
            [sg.Text("----------------------------------------------------------------------------------------",text_color="#d0e4f3",font=font)],
            [sg.FileBrowse("製造実工数",font=font) ,sg.InputText(font=font,default_text=ActualTimepath,background_color="#02497c",text_color="#ffffff",size=(55,1))],
            [sg.Button("実行_Auto",button_color=SetButtonColor),sg.Button("実行_Manu",button_color=SetButtonColor),sg.Button("閉じる",button_color="#b83811")],
            [sg.Text("----------------------------------------------------------------------------------------",text_color="#d0e4f3",font=font)]]


    # ----------------------------------------------------------------------------------------------------------------------------------------------------
    # ウィンドウの生成
    window = sg.Window("MS9740B_進捗確認Tool", layout)
    # イベントループ
    while True:
        event, values = window.read()
        if type(values) is dict:
            # print(type(values))
            # コンポーネントに入ったデータの代入
            SGExcelPathInput = values[0] 
            MSExcelPathInput = values[1]
            HTMLOutputPath = values[2]
            AllSetModeInput = values[3]
            OneSetModeInput = values[4]
            SetSerialInput = values[5]
            SetKougakubuPath= values[6]
            SetKanryouPath= values[7]
            SetActualTimepath = values[8]

        # 閉じるボタン押下処理
        if event == sg.WIN_CLOSED or event == "閉じる":
            # 設定パスをConfig.iniに保存
            if SGExcelPathInput != None:
                config.set("Path","SGExcelPath",SGExcelPathInput)
                config.set("Path","MSExcelPath",MSExcelPathInput)
                config.set("Path","htmloutpath",HTMLOutputPath)
                config.set("DefaultSet","ProvisionalNumber",SetSerialInput)
                config.set("Path","kougakubupath",SetKougakubuPath)
                config.set("Path","kanryoupath",SetKanryouPath)
                config.set("Path","actualtimepath",SetActualTimepath)
                
                with open(ConfigIniPath, "w",encoding="utf-8") as file:
                    config.write(file)
                # jsonファイル保存
                jsonPath = os.path.join(os.path.dirname(os.path.abspath(__file__)) ,"Config.json")
                inijson(jsonPath,ConfigIniPath)
                # ウィンドウ閉じる
                break
            else:
                break
        elif event == "更新":
            print("-----------------------------------------------------------------------------------------")
            print ("進捗情報に必要なデータを更新します。Excel進捗表から収集し「1」～「4」実行で完了です。")
            print("全仮番取得：" + str(AllSetModeInput) + "　仕掛品取得：" + str(OneSetModeInput) + "　モードで取得中")
            print("-----------------------------------------------------------------------------------------")
            # print("更新", SGExcelPathInput)
            # 計画用管理シートの設定
            SGExcelPath = SGExcelPathInput
            SGGetSheetName = "工程進捗表（最新）"
            SGStartRowSet = 5
            SGDataInfoStartRowSet = 6
            # ファイル存在確認
            if os.path.exists(SGExcelPath) == False:
                print("指定したファイルパスは存在しません。Path：" + SGExcelPath)
                break
            # 進捗管理表をコピー
            SGExcelCopyPath = os.path.join(os.path.dirname(os.path.abspath(__file__)) ,"ExcelCopyFile")
            SGExcelCopyPath = os.path.join(SGExcelCopyPath,"MS9740A工程進捗表.xlsx")
            if os.path.exists(SGExcelCopyPath) == True:
                os.remove(SGExcelCopyPath)
            shutil.copy(SGExcelPath,SGExcelCopyPath)
            # スクリプトファイル同階層にパス設定
            HTMLPath = os.path.join(os.path.dirname(os.path.abspath(__file__)) ,"ScheduleProgressGet.html")
            SGvalueDatPath = os.path.join(os.path.dirname(os.path.abspath(__file__)) ,"SGvalueDat.pkl")
            # ヘッダ検索セルヒット条件文字列
            MainMachString = "本体仮Ｎｏ"
            listDataMachString = ["本体シリアル番号","受注番号","ｴﾝﾄﾞﾕｰｻﾞｰ"]
            # 処理時間測定用
            start = time.time()
            # ワークブックを読取り専用で開く
            book = excel.load_workbook(SGExcelCopyPath,data_only=True,read_only=True)
            # ワークシートを取り出す
            sheet = book[SGGetSheetName]
            # 処理時間表示
            elapsed_time = time.time() - start
            # print ("SG用進捗管理通過","elapsed_time:{0}".format(elapsed_time) + "[sec]")
            print ("「1」MS9740A工程進捗表のデータ取得が完了しました。")
            print ("     経過時間：{:.1f}".format(elapsed_time) + "[秒]")
            print("-----------------------------------------------------------------------------------------")
            # 辞書型の変数を初期化
            SGvalueDat = {} 
            # 多次元辞書型にデータを格納
            SGvalueDat = CreateDic(sheet,MainMachString,listDataMachString,SGStartRowSet,SGDataInfoStartRowSet)
            book.close()
            # 処理時間表示
            elapsed_time = time.time() - start
            # print ("SG用データ格納通過","elapsed_time:{0}".format(elapsed_time) + "[sec]")
            print ("「2」データの格納が完了しました。")
            print ("     経過時間：{:.1f}".format(elapsed_time) + "[秒]")
            print("-----------------------------------------------------------------------------------------")
            # 製造用管理シートの設定
            MSExcelPath = MSExcelPathInput
            MSGetSheetName = "日程表（最新）"
            MSStartRowSet = 4
            MSDataInfoStartRowSet = 5
            # MSDataInfoStartRowSet = 1500
            # ファイル存在確認
            if os.path.exists(MSExcelPath) == False:
                print("指定したファイルパスは存在しません。Path：" + MSExcelPath)
                break
            # 進捗管理表をコピー
            MSExcelCopyPath = os.path.join(os.path.dirname(os.path.abspath(__file__)) ,"ExcelCopyFile")
            MSExcelCopyPath = os.path.join(MSExcelCopyPath,"MS9740A チェックシート進捗管理版(原本).xlsm")
            if os.path.exists(MSExcelCopyPath) == True:
                os.remove(MSExcelCopyPath)
            shutil.copy(MSExcelPath,MSExcelCopyPath)
            MSdicColDay = {} 
            # ワークブックを読取り専用で開く
            MSbook = excel.load_workbook(MSExcelCopyPath,data_only=True,read_only=True)
            # ワークシートを取り出す
            MSsheet = MSbook[MSGetSheetName]
            # 辞書型にColumn:日付_yyyy/mm/ddを代入
            MSdicColDay = CreateDicDay(MSsheet,MSStartRowSet)
            # 処理時間表示
            elapsed_time = time.time() - start
            print ("「3」MS9740A チェックシート進捗管理版(原本)のデータ取得が完了しました。")
            print ("     経過時間：{:.1f}".format(elapsed_time) + "[秒]")
            print ("-----------------------------------------------------------------------------------------")
            # シート内全データを取得
            for row in MSsheet.iter_rows(min_row=MSDataInfoStartRowSet):
                # 指定行の列、セルを取得
                strUpdate = ""
                for index , cell in enumerate(row,start=1):
                    # セルの値が期待した値だった場合の処理
                    if cell.value is not None and cell.value is not 0:
                        # その値が辞書型に入っている場合の処理
                        if index == 10:
                            if str(cell.value) in SGvalueDat.keys():
                                strUpdate = str(cell.value)
                                # SGvalueDat[strUpdate].update({"":""})
                    # 進捗表のデータを辞書型配列に更新させる＿最新進捗状況を取得
                    if strUpdate != "":
                        if str(cell.value) == "指図":
                            SGvalueDat = UPDateDicData(MSdicColDay,SGvalueDat,strUpdate,"指図",cell.column,cell.fill.patternType,cell.fill.fgColor.value,cell.fill.bgColor.value)
                        elif str(cell.value) == "P":
                            SGvalueDat = UPDateDicData(MSdicColDay,SGvalueDat,strUpdate,"P",cell.column,cell.fill.patternType,cell.fill.fgColor.value,cell.fill.bgColor.value)
                        elif str(cell.value) == "組":
                            SGvalueDat = UPDateDicData(MSdicColDay,SGvalueDat,strUpdate,"組",cell.column,cell.fill.patternType,cell.fill.fgColor.value,cell.fill.bgColor.value)
                        elif str(cell.value) == "立":
                            SGvalueDat = UPDateDicData(MSdicColDay,SGvalueDat,strUpdate,"立",cell.column,cell.fill.patternType,cell.fill.fgColor.value,cell.fill.bgColor.value)
                        elif str(cell.value) == "ｋ3":
                            SGvalueDat = UPDateDicData(MSdicColDay,SGvalueDat,strUpdate,"ｋ3",cell.column,cell.fill.patternType,cell.fill.fgColor.value,cell.fill.bgColor.value)
                        elif str(cell.value) == "高前":
                            SGvalueDat = UPDateDicData(MSdicColDay,SGvalueDat,strUpdate,"高前",cell.column,cell.fill.patternType,cell.fill.fgColor.value,cell.fill.bgColor.value)
                        elif str(cell.value) == "低":
                            SGvalueDat = UPDateDicData(MSdicColDay,SGvalueDat,strUpdate,"低",cell.column,cell.fill.patternType,cell.fill.fgColor.value,cell.fill.bgColor.value)
                        elif str(cell.value) == "高温":
                            SGvalueDat = UPDateDicData(MSdicColDay,SGvalueDat,strUpdate,"高温",cell.column,cell.fill.patternType,cell.fill.fgColor.value,cell.fill.bgColor.value)
                        elif str(cell.value) == "ｋ4":
                            SGvalueDat = UPDateDicData(MSdicColDay,SGvalueDat,strUpdate,"ｋ4",cell.column,cell.fill.patternType,cell.fill.fgColor.value,cell.fill.bgColor.value)
                        elif str(cell.value) == "出準":
                            SGvalueDat = UPDateDicData(MSdicColDay,SGvalueDat,strUpdate,"出準",cell.column,cell.fill.patternType,cell.fill.fgColor.value,cell.fill.bgColor.value)
                        elif str(cell.value) == "発送":
                            SGvalueDat = UPDateDicData(MSdicColDay,SGvalueDat,strUpdate,"発送",cell.column,cell.fill.patternType,cell.fill.fgColor.value,cell.fill.bgColor.value)
            MSbook.close()
            # 処理時間表示
            elapsed_time = time.time() - start
            print ("「4」全ての更新が完了しました。")
            print ("     経過時間：{:.1f}".format(elapsed_time) + "[秒]")
            print("-----------------------------------------------------------------------------------------")
            SasizuDayList = []
            HassouDayList = []
            KariNo = []
            # SecondKeys = ["本体シリアル番号","受注番号","ｴﾝﾄﾞﾕｰｻﾞｰ","指図_日付","指図_進捗","P_日付","P_進捗","組_日付","組_進捗","立_日付","立_進捗","ｋ3_日付","ｋ3_進捗","高前_日付","高前_進捗","低_日付","低_進捗","高温_日付","高温_進捗","ｋ4_日付","ｋ4_進捗","出準_日付","出準_進捗","発送_日付","発送_進捗"]
            # 吸い上げたデータをデータフレームに格納する
            for FirstKey in SGvalueDat.keys():
                # 仮番号：4300以上に対して処理する
                if isint(FirstKey):
                    # if int(FirstKey) >= 4300:
                    if int(FirstKey) >= int(SetSerialInput):
                        if "出準_進捗" in SGvalueDat[FirstKey].keys():
                            # 更に出荷準備が未完了のもを抽出
                            if  OneSetModeInput == True and  SGvalueDat[FirstKey]["出準_進捗"] == "未":
                                KariNo.append(str(FirstKey))
                                # print("仮番号",FirstKey)
                                for SecondKey in SGvalueDat[FirstKey].keys():
                                    if SecondKey == "指図_日付":
                                        SasizuDayList.append(SGvalueDat[FirstKey][SecondKey])
                                    if SecondKey == "発送_日付":
                                        HassouDayList.append(SGvalueDat[FirstKey][SecondKey])
                            elif AllSetModeInput == True:
                                KariNo.append(str(FirstKey))
                                # print("仮番号",FirstKey)
                                for SecondKey in SGvalueDat[FirstKey].keys():
                                    if SecondKey == "指図_日付":
                                        SasizuDayList.append(SGvalueDat[FirstKey][SecondKey])
                                    if SecondKey == "発送_日付":
                                        HassouDayList.append(SGvalueDat[FirstKey][SecondKey])
                                    # print(SecondKey,SGvalueDat[FirstKey][SecondKey])
            # 日付条件の設定
            strdt = dt.strptime(min(SasizuDayList), '%Y/%m/%d')  # 開始日
            enddt = dt.strptime(max(HassouDayList), '%Y/%m/%d')  # 終了日
            # 日付差の日数を算出（リストに最終日も含めたいので、＋１しています）
            days_num = (enddt - strdt).days + 1  # （参考）括弧の部分はtimedelta型のオブジェクトになります
            # シンプルにforとappendを使用した場合
            datelist = []
            for i in range(days_num):
                daydat = strdt + timedelta(days=i)
                datelist.append(str(daydat.strftime("%Y/%m/%d")))
            df = pd.DataFrame(index=KariNo,columns=["本体シリアル番号","受注番号","ｴﾝﾄﾞﾕｰｻﾞｰ"] + datelist)
            # df = pd.DataFrame(columns=[datelist])
            # SecondKeys = ["本体シリアル番号","受注番号","ｴﾝﾄﾞﾕｰｻﾞｰ","指図_日付","指図_進捗","P_日付","P_進捗","組_日付","組_進捗","立_日付","立_進捗","ｋ3_日付","ｋ3_進捗","高前_日付","高前_進捗","低_日付","低_進捗","高温_日付","高温_進捗","ｋ4_日付","ｋ4_進捗","出準_日付","出準_進捗","発送_日付","発送_進捗"]
            # 吸い上げたデータをデータフレームに格納する
            for FirstKey in SGvalueDat.keys():
                # 仮番号：4300以上に対して処理する
                if isint(FirstKey):
                    if int(FirstKey) >= int(SetSerialInput):
                        if "出準_進捗" in SGvalueDat[FirstKey].keys():
                            # 更に出荷準備が未完了のもを抽出
                            if OneSetModeInput == True and SGvalueDat[FirstKey]["出準_進捗"] == "未":
                                # print("仮番号",FirstKey)
                                # df = df.index(FirstKey)
                                for SecondKey in SGvalueDat[FirstKey].keys():
                                    # print(SGvalueDat[FirstKey][SecondKey])
                                    if SecondKey == "本体シリアル番号":
                                        df.loc[FirstKey,SecondKey] = str(SGvalueDat[FirstKey][SecondKey])
                                    if SecondKey == "受注番号":
                                        df.loc[FirstKey,SecondKey] = str(SGvalueDat[FirstKey][SecondKey])
                                    if SecondKey == "ｴﾝﾄﾞﾕｰｻﾞｰ":
                                        df.loc[FirstKey,SecondKey] = str(SGvalueDat[FirstKey][SecondKey])
                                    if SecondKey == "指図_日付":
                                        df.loc[FirstKey,SGvalueDat[FirstKey][SecondKey]] = "指図"
                                    if SecondKey == "P_日付":
                                        df.loc[FirstKey,SGvalueDat[FirstKey][SecondKey]] = "P"
                                    if SecondKey == "組_日付":
                                        df.loc[FirstKey,SGvalueDat[FirstKey][SecondKey]] = "組立"
                                    if SecondKey == "立_日付":
                                        df.loc[FirstKey,SGvalueDat[FirstKey][SecondKey]] = "立上"
                                    if SecondKey == "ｋ3_日付":
                                        df.loc[FirstKey,SGvalueDat[FirstKey][SecondKey]] = "ｋ3"
                                    if SecondKey == "高前_日付":
                                        df.loc[FirstKey,SGvalueDat[FirstKey][SecondKey]] = "高前"
                                    if SecondKey == "低_日付":
                                        df.loc[FirstKey,SGvalueDat[FirstKey][SecondKey]] = "低温"
                                    if SecondKey == "高温_日付":
                                        df.loc[FirstKey,SGvalueDat[FirstKey][SecondKey]] = "高温"
                                    if SecondKey == "ｋ4_日付":
                                        df.loc[FirstKey,SGvalueDat[FirstKey][SecondKey]] = "ｋ4"
                                    if SecondKey == "出準_日付":
                                        df.loc[FirstKey,SGvalueDat[FirstKey][SecondKey]] = "出準"
                                    if SecondKey == "発送_日付":
                                        df.loc[FirstKey,SGvalueDat[FirstKey][SecondKey]] = "発送"
                            elif AllSetModeInput == True:
                                # print("仮番号",FirstKey)
                                # df = df.index(FirstKey)
                                for SecondKey in SGvalueDat[FirstKey].keys():
                                    # print(SGvalueDat[FirstKey][SecondKey])
                                    if SecondKey == "本体シリアル番号":
                                        df.loc[FirstKey,SecondKey] = str(SGvalueDat[FirstKey][SecondKey])
                                    if SecondKey == "受注番号":
                                        df.loc[FirstKey,SecondKey] = str(SGvalueDat[FirstKey][SecondKey])
                                    if SecondKey == "ｴﾝﾄﾞﾕｰｻﾞｰ":
                                        df.loc[FirstKey,SecondKey] = str(SGvalueDat[FirstKey][SecondKey])
                                    if SecondKey == "指図_日付":
                                        df.loc[FirstKey,SGvalueDat[FirstKey][SecondKey]] = "指図"
                                    if SecondKey == "P_日付":
                                        df.loc[FirstKey,SGvalueDat[FirstKey][SecondKey]] = "P"
                                    if SecondKey == "組_日付":
                                        df.loc[FirstKey,SGvalueDat[FirstKey][SecondKey]] = "組立"
                                    if SecondKey == "立_日付":
                                        df.loc[FirstKey,SGvalueDat[FirstKey][SecondKey]] = "立上"
                                    if SecondKey == "ｋ3_日付":
                                        df.loc[FirstKey,SGvalueDat[FirstKey][SecondKey]] = "ｋ3"
                                    if SecondKey == "高前_日付":
                                        df.loc[FirstKey,SGvalueDat[FirstKey][SecondKey]] = "高前"
                                    if SecondKey == "低_日付":
                                        df.loc[FirstKey,SGvalueDat[FirstKey][SecondKey]] = "低温"
                                    if SecondKey == "高温_日付":
                                        df.loc[FirstKey,SGvalueDat[FirstKey][SecondKey]] = "高温"
                                    if SecondKey == "ｋ4_日付":
                                        df.loc[FirstKey,SGvalueDat[FirstKey][SecondKey]] = "ｋ4"
                                    if SecondKey == "出準_日付":
                                        df.loc[FirstKey,SGvalueDat[FirstKey][SecondKey]] = "出準"
                                    if SecondKey == "発送_日付":
                                        df.loc[FirstKey,SGvalueDat[FirstKey][SecondKey]] = "発送"
            # データフレームデータの欠損Nanを削除
            df = df.fillna("")
            # データフレームをhtmlで保存する
            # df.to_html(r"X:\Python\python自動処理の教科書\OkumuraStudy\ScheduleProgressGet.html",justify="left",encoding="utf-8")
            html = df.to_html()
            text_file = open(HTMLPath,"w")
            text_file.write(html)
            text_file.close()
            # 辞書をファイルとして保存
            with open(SGvalueDatPath,"wb") as f:
                pickle.dump(SGvalueDat, f)
            print ("進捗表を表示する場合は「出力」を押下して下さい。")
            print("-----------------------------------------------------------------------------------------")
        elif event == "出力":
            # print("更新", MSExcelPathInput)
            # HTMLファイルをブラウザで開く
            HTMLPath = os.path.join(os.path.dirname(os.path.abspath(__file__)) ,"ScheduleProgressGet.html")
            StyleHTMLPath = os.path.join(os.path.dirname(os.path.abspath(__file__)) ,"ScheduleProgressHTML.html")
            SGvalueDatPath = os.path.join(os.path.dirname(os.path.abspath(__file__)) ,"SGvalueDat.pkl")
            CSSPath = os.path.join(os.path.dirname(os.path.abspath(__file__)) ,"ScheduleProgressGet.css")
            JsPath = os.path.join(os.path.dirname(os.path.abspath(__file__)) ,"ScheduleProgressHTML.js")
            # ScheduleProgressGet.pyで生成されたデータを読み込む
            with open(SGvalueDatPath,"rb") as f:
                SGvalueDat = pickle.load(f)
            # htmlファイルを読み込む
            with open(HTMLPath, mode="rt") as f:
                html = f.read()
            soup = BeautifulSoup(html, "html.parser")
            # cssリンクタグを作る
            nt = soup.new_tag("link", rel="stylesheet",href="ScheduleProgressGet.css")
            # table の子のインデックス0の位置 (つまり、先頭) に挿入する。
            soup.table.insert(0, nt)
            # cssリンクタグを作る
            nt = soup.new_tag("script", src="ScheduleProgressHTML.js")
            # table の子のインデックス0の位置 (つまり、先頭) に挿入する。
            soup.table.insert(0, nt)
            nt = soup.new_tag("button",attrs={"class": "fixed_btn","type":"button","onclick":"ExcelWrite()"})
            nt.string = "ExcelSave"
            # table の子のインデックス0の位置 (つまり、先頭) に挿入する。
            soup.table.insert(0, nt)
            nt = soup.new_tag("title")
            nt.string = "MS9740B_ScheduleProgress"
            # table の子のインデックス0の位置 (つまり、先頭) に挿入する。
            soup.table.insert(0, nt)
            # テーブルの取得
            table = soup.find("table")
            table.attrs["id"] = "data"
            # theadの解析
            thead = table.find("thead")
            ths = thead.tr.find_all("th")
            # 曜日変換用辞書
            daydict = {"Mon":"月","Tue":"火","Wed":"水","Thu":"木","Fri":"金","Sat":"土","Sun":"日"}
            for i,th in enumerate(ths):
                # 5列目から日付表示処理
                if th.text == "本体シリアル番号":
                    th.attrs["class"] = "Sn"
                if th.text == "受注番号":
                    th.attrs["class"] = "OrderNumber"
                if th.text == "ｴﾝﾄﾞﾕｰｻﾞｰ":
                    th.attrs["class"] = "EndUser"
                if i > 3 :
                    # 日付を短縮表示、曜日変換
                    day = dt.strptime(th.text,"%Y/%m/%d")
                    th.string.replace_with(th.text[5:],daydict[day.strftime("%a")]) 
                    # 日付タグが今日の日付だった場合セル色を変える
                    if day.date() == days.today():
                        th.attrs["class"] = "today"
            # tbodyの解析
            tbody = table.find("tbody")
            trs = tbody.find_all("tr")
            kancalor = "#599ac9"
            for tr in trs:
                for i,th in enumerate(tr.find_all("th")):
                    for i,td in enumerate(tr.find_all("td")):
                        # タグにクラス属性追加
                        if td.text == "指図" and SGvalueDat[th.text]["指図_進捗"] == "完":
                            td.attrs["bgcolor"] = kancalor
                        elif td.text == "P" and SGvalueDat[th.text]["P_進捗"] == "完":
                            td.attrs["bgcolor"] = kancalor
                        elif td.text == "組立" and SGvalueDat[th.text]["組_進捗"] == "完":
                            td.attrs["bgcolor"] = kancalor
                        elif td.text == "立上" and SGvalueDat[th.text]["立_進捗"] == "完":
                            td.attrs["bgcolor"] = kancalor
                        elif td.text == "ｋ3" and SGvalueDat[th.text]["ｋ3_進捗"] == "完":
                            td.attrs["bgcolor"] = kancalor
                        elif td.text == "高前" and SGvalueDat[th.text]["高前_進捗"] == "完":
                            td.attrs["bgcolor"] = kancalor
                        elif td.text == "低温" and SGvalueDat[th.text]["低_進捗"] == "完":
                            td.attrs["bgcolor"] = kancalor
                        elif td.text == "高温" and SGvalueDat[th.text]["高温_進捗"] == "完":
                            td.attrs["bgcolor"] = kancalor
                        elif td.text == "ｋ4" and SGvalueDat[th.text]["ｋ4_進捗"] == "完":
                            td.attrs["bgcolor"] = kancalor
                        elif td.text == "出準" and SGvalueDat[th.text]["出準_進捗"] == "完":
                            td.attrs["bgcolor"] = kancalor
                        elif td.text == "発送" and SGvalueDat[th.text]["発送_進捗"] == "完":
                            td.attrs["bgcolor"] = kancalor
                        else:
                            td.attrs["bgcolor"] = "#d9e2e6"
            # htmlファイルを保存する
            with open(StyleHTMLPath, "w") as file: 
                file.write(str(soup.prettify())) 
            jsonPath = os.path.join(os.path.dirname(os.path.abspath(__file__)) ,"Config.json")
            # Webで開く
            uri = "file:" + StyleHTMLPath
            webbrowser.open(uri,autoraise=True)
            shutil.copy(StyleHTMLPath,os.path.join(HtmlOutPath ,"ScheduleProgressHTML.html"))
            shutil.copy(CSSPath,os.path.join(HtmlOutPath ,"ScheduleProgressGet.css"))
            shutil.copy(JsPath,os.path.join(HtmlOutPath ,"ScheduleProgressHTML.js"))
            shutil.copy(jsonPath,os.path.join(HtmlOutPath ,"Config.json"))
        elif event == "光学部開く":
            pro = subprocess.Popen(SetKougakubuPath,shell=True)
            print("-----------------------------------------------------------------------------------------")
            print("MS9740A 光学部進捗管理版(原本).xlsmを開いています")
            print("-----------------------------------------------------------------------------------------")
        elif event == "完了表開く":
            pro = subprocess.Popen(SetKanryouPath,shell=True)
            print("日程管理Daily_MS9740B.xlsxを開いています")
            print("-----------------------------------------------------------------------------------------")
        elif event == "R/3(002)実行":
            print("R/3(002)を実行しています")
            # pro = subprocess.Popen(r"\\Fks-file-005\disk4\Project13\工程管理\MS9740A_進捗管理表\MS9740B生産向上WG\日程管理\FY21_日程管理Daily_MS9740B.xlsx",shell=True)
            # R3を開く
            pro = subprocess.Popen(r"C:\Program Files (x86)\SAP\FrontEnd\SAPgui\saplogon.exe",shell=True)
            pgui.FAILSAFE = True
            # R3が開くまで待つ
            while pgui.locateOnScreen(r".\img\R3LogIn.png" , confidence=0.9) is None:
                time.sleep(1)
            pgui.hotkey("enter")
            # ログインが開くまで待つ
            while pgui.locateOnScreen(r".\img\R3LogIn_Pass3.png" , confidence=0.9) is None:
                time.sleep(1)
            clip.copy("rt002")
            pgui.hotkey("ctrl","v")
            pgui.press("tab")
            clip.copy("pass002")
            pgui.hotkey("ctrl","v")
            pgui.press("enter")
            # R3が開くまで待つ
            while pgui.locateOnScreen(r".\img\R3Main.png" , confidence=0.9) is None:
                time.sleep(1)
            time.sleep(2)
            pro = subprocess.Popen(r".\vbs\MS9740B.vbs",shell=True)
            # R3が開くまで待つ
            while pgui.locateOnScreen(r".\img\vbs9740b.png" , confidence=0.9) is None:
                time.sleep(1)
            pgui.hotkey("alt","o")
            time.sleep(1)
            pgui.hotkey("alt","tab")
            input("Press enter to start operation...")
            time.sleep(1)
            pgui.hotkey("alt","tab")
            time.sleep(1)
            pgui.hotkey("F3")
            time.sleep(1)
            pgui.hotkey("F3")
            time.sleep(1)
            # R3が開くまで待つ
            while pgui.locateOnScreen(r".\img\R3Main.png" , confidence=0.9) is None:
                time.sleep(1)
            time.sleep(2)
            pro = subprocess.Popen(r".\vbs\光学部.vbs",shell=True)
            # R3が開くまで待つ
            while pgui.locateOnScreen(r".\img\vbs9740b.png" , confidence=0.9) is None:
                time.sleep(1)
            pgui.hotkey("alt","o")
            time.sleep(1)
            pgui.hotkey("alt","tab")
            print("-----------------------------------------------------------------------------------------")
        elif event == "R/3(003)実行":
            print("R/3(003)を実行しています")
            # pro = subprocess.Popen(r"\\Fks-file-005\disk4\Project13\工程管理\MS9740A_進捗管理表\MS9740B生産向上WG\日程管理\FY21_日程管理Daily_MS9740B.xlsx",shell=True)
            # R3を開く
            pro = subprocess.Popen(r"C:\Program Files (x86)\SAP\FrontEnd\SAPgui\saplogon.exe",shell=True)
            pgui.FAILSAFE = True
            # R3が開くまで待つ
            while pgui.locateOnScreen(r".\img\R3LogIn.png" , confidence=0.9) is None:
                time.sleep(1)
            pgui.hotkey("enter")
            # ログインが開くまで待つ
            while pgui.locateOnScreen(r".\img\R3LogIn_Pass3.png" , confidence=0.9) is None:
                time.sleep(1)
            clip.copy("rt003")
            pgui.hotkey("ctrl","v")
            pgui.press("tab")
            clip.copy("pass003")
            pgui.hotkey("ctrl","v")
            pgui.press("enter")
            # R3が開くまで待つ
            while pgui.locateOnScreen(r".\img\R3Main.png" , confidence=0.9) is None:
                time.sleep(1)
            time.sleep(2)
            pro = subprocess.Popen(r".\vbs\MS9740B.vbs",shell=True)
            # R3が開くまで待つ
            while pgui.locateOnScreen(r".\img\vbs9740b.png" , confidence=0.9) is None:
                time.sleep(1)
            pgui.hotkey("alt","o")
            time.sleep(1)
            pgui.hotkey("alt","tab")
            input("Press enter to start operation...")
            time.sleep(1)
            pgui.hotkey("alt","tab")
            time.sleep(1)
            pgui.hotkey("F3")
            time.sleep(1)
            pgui.hotkey("F3")
            time.sleep(1)
            # R3が開くまで待つ
            while pgui.locateOnScreen(r".\img\R3Main.png" , confidence=0.9) is None:
                time.sleep(1)
            time.sleep(2)
            pro = subprocess.Popen(r".\vbs\光学部.vbs",shell=True)
            # R3が開くまで待つ
            while pgui.locateOnScreen(r".\img\vbs9740b.png" , confidence=0.9) is None:
                time.sleep(1)
            pgui.hotkey("alt","o")
            time.sleep(1)
            pgui.hotkey("alt","tab")
            print("-----------------------------------------------------------------------------------------")
        elif event == "実行_Auto":
            FirstDate = get_first_date(datetime.datetime.today())
            strFirstDate = FirstDate.strftime("%Y/%m/%d")
            print("-----------------------------------------------------------------------------------------")
            print ("製造集計Toolから実績工数ファイルを出力します。開始日付：" + strFirstDate)
            print("-----------------------------------------------------------------------------------------")
            # 製造実績集計ツールを開く
            pgui.press("win")
            # 製造実績集計ツールが開くまで待つ
            while pgui.locateOnScreen(r".\img\SeizoTool.png" , confidence=0.9) is None:
                time.sleep(1)
            # 現在位置を取得
            loc = pgui.position()
            offset=(0, 0)
            # 製造実績集計ツールボタンの位置を特定
            img_x, img_y = pgui.locateCenterOnScreen(r".\img\SeizoTool.png", grayscale=True, confidence=0.9)
            # オフセット分を計算
            img_x = img_x + offset[0]
            img_y = img_y + offset[1]
            # 対象位置へ移動
            pgui.moveTo(img_x,img_y,1)
            pgui.click(img_x,img_y)
            # 製造実績集計ツールが開くまで待つ
            while pgui.locateOnScreen(r".\img\OutButton.png" , confidence=0.9) is None:
                time.sleep(1)
            # 現在位置を取得
            loc = pgui.position()
            offset=(0, 0)
            # 製造実績集計ツールボタンの位置を特定
            img_x, img_y = pgui.locateCenterOnScreen(r".\img\OutButton.png", grayscale=True, confidence=0.9)
            # オフセット分を計算
            img_x = img_x + offset[0]
            img_y = img_y + offset[1]
            # 対象位置へ移動
            pgui.moveTo(img_x,img_y,1)
            pgui.click(img_x,img_y)
            # 製造実績集計ツールが開くまで待つ
            while pgui.locateOnScreen(r".\img\SeizoToo2.png" , confidence=0.9) is None:
                time.sleep(1)
            
            # 仕掛工数情報出力の親ハンドルを取得
            parent_handle = ctypes.windll.user32.FindWindowW(0, "仕掛工数情報出力")
            # 仕掛工数情報出力をアクティブにする
            ctypes.windll.user32.SetForegroundWindow(parent_handle)
            
            # 開始日付
            pgui.press("tab",3,interval=0.3)
            pgui.typewrite(strFirstDate)
            # 作業部門
            pgui.press("tab",5,interval=0.3)
            pgui.typewrite("GMMS11")
            # 完了
            pgui.press("tab",4,interval=0.3)
            time.sleep(0.3)
            pgui.press("space")
            # 現在位置を取得
            loc = pgui.position()
            offset=(0, 0)
            # 製造実績集計ツールボタンの位置を特定
            img_x, img_y = pgui.locateCenterOnScreen(r".\img\StartButton.png", grayscale=True, confidence=0.9)
            # オフセット分を計算
            img_x = img_x + offset[0]
            img_y = img_y + offset[1]
            # 対象位置へ移動
            pgui.moveTo(img_x,img_y,1)
            pgui.click(img_x,img_y)
            # 製造実績集計ツールが開くまで待つ
            while pgui.locateOnScreen(r".\img\Executing.png" , confidence=0.9) is not None:
                time.sleep(1)
            
            parent_handle = ctypes.windll.user32.FindWindowW(0, "仕掛工数情報出力")
            # 仕掛工数情報出力をアクティブにする
            ctypes.windll.user32.SetForegroundWindow(parent_handle)            
            
            
            # 現在位置を取得
            loc = pgui.position()
            offset=(0, 0)
            # 製造実績集計ツールボタンの位置を特定
            img_x, img_y = pgui.locateCenterOnScreen(r".\img\ExcelButton.png", grayscale=True, confidence=0.9)
            # オフセット分を計算
            img_x = img_x + offset[0]
            img_y = img_y + offset[1]
            # 対象位置へ移動
            pgui.moveTo(img_x,img_y,1)
            pgui.click(img_x,img_y)
            # 製造実績集計ツールが開くまで待つ
            while pgui.locateOnScreen(r".\img\PathIn.png" , confidence=0.9) is None:
                time.sleep(1)
            
            parent_handle = ctypes.windll.user32.FindWindowW(0, "出力先のファイル名を選択してください")
            # 出力先のファイル名を選択してくださいをアクティブにする
            ctypes.windll.user32.SetForegroundWindow(parent_handle)
            
            
            pgui.press("tab",6,interval=0.3)
            pgui.press("enter")
            clip.copy(SetActualTimepath)
            pgui.hotkey("ctrl","v")
            # pgui.typewrite(SetActualTimepath)
            time.sleep(0.3)
            pgui.press("enter")
            time.sleep(0.3)
            pgui.hotkey("alt","s")
            print ("処理が完了しました。保存先は下記を参照。" )
            print (SetActualTimepath)
            print("-----------------------------------------------------------------------------------------")
        elif event == "実行_Manu":            
            FirstDate = get_first_date(datetime.datetime.today())
            strFirstDate = FirstDate.strftime("%Y/%m/%d")
            print("-----------------------------------------------------------------------------------------")
            print ("製造集計Toolから実績工数ファイルを出力します。開始日付：" + strFirstDate)
            print("-----------------------------------------------------------------------------------------")
            # 製造実績集計ツールを開く
            pgui.press("win")
            # 製造実績集計ツールが開くまで待つ
            while pgui.locateOnScreen(r".\img\SeizoTool.png" , confidence=0.9) is None:
                time.sleep(1)
            # 現在位置を取得
            loc = pgui.position()
            offset=(0, 0)
            # 製造実績集計ツールボタンの位置を特定
            img_x, img_y = pgui.locateCenterOnScreen(r".\img\SeizoTool.png", grayscale=True, confidence=0.9)
            # オフセット分を計算
            img_x = img_x + offset[0]
            img_y = img_y + offset[1]
            # 対象位置へ移動
            pgui.moveTo(img_x,img_y,1)
            pgui.click(img_x,img_y)
            # 製造実績集計ツールが開くまで待つ
            while pgui.locateOnScreen(r".\img\OutButton.png" , confidence=0.9) is None:
                time.sleep(1)
            # 現在位置を取得
            loc = pgui.position()
            offset=(0, 0)
            # 製造実績集計ツールボタンの位置を特定
            img_x, img_y = pgui.locateCenterOnScreen(r".\img\OutButton.png", grayscale=True, confidence=0.9)
            # オフセット分を計算
            img_x = img_x + offset[0]
            img_y = img_y + offset[1]
            # 対象位置へ移動
            pgui.moveTo(img_x,img_y,1)
            pgui.click(img_x,img_y)
            # 製造実績集計ツールが開くまで待つ
            while pgui.locateOnScreen(r".\img\SeizoToo2.png" , confidence=0.9) is None:
                time.sleep(1)
            # 開始日付
            input("Press enter to start operation...")
            time.sleep(1) 
            
            # 0000000000031204
            # 仕掛工数情報出力の親ハンドルを取得
            parent_handle = ctypes.windll.user32.FindWindowW(0, "仕掛工数情報出力")
            # 仕掛工数情報出力をアクティブにする
            ctypes.windll.user32.SetForegroundWindow(parent_handle)
            
            pgui.press("tab",3,interval=0.3)
            pgui.typewrite(strFirstDate)
            # 作業部門
            pgui.press("tab",5,interval=0.3)
            pgui.typewrite("GMMS11")
            # 完了
            pgui.press("tab",4,interval=0.3)
            time.sleep(0.3)
            pgui.press("space")
            # 現在位置を取得
            loc = pgui.position()
            offset=(0, 0)
            # 製造実績集計ツールボタンの位置を特定
            img_x, img_y = pgui.locateCenterOnScreen(r".\img\StartButton.png", grayscale=True, confidence=0.9)
            # オフセット分を計算
            img_x = img_x + offset[0]
            img_y = img_y + offset[1]
            # 対象位置へ移動
            pgui.moveTo(img_x,img_y,1)
            pgui.click(img_x,img_y)
            # 製造実績集計ツールが開くまで待つ
            while pgui.locateOnScreen(r".\img\Executing.png" , confidence=0.9) is not None:
                time.sleep(1)
            input("Press enter to start operation...")
            time.sleep(1)
            
            # 0000000000031204
            parent_handle = ctypes.windll.user32.FindWindowW(0, "仕掛工数情報出力")
            # 仕掛工数情報出力をアクティブにする
            ctypes.windll.user32.SetForegroundWindow(parent_handle)
            
            # 現在位置を取得
            loc = pgui.position()
            offset=(0, 0)
            # 製造実績集計ツールボタンの位置を特定
            img_x, img_y = pgui.locateCenterOnScreen(r".\img\ExcelButton.png", grayscale=True, confidence=0.9)
            # オフセット分を計算
            img_x = img_x + offset[0]
            img_y = img_y + offset[1]
            # 対象位置へ移動
            pgui.moveTo(img_x,img_y,1)
            pgui.click(img_x,img_y)
            # 製造実績集計ツールが開くまで待つ
            while pgui.locateOnScreen(r".\img\PathIn.png" , confidence=0.9) is None:
                time.sleep(1)
            input("Press enter to start operation...")
            time.sleep(1)
            
            # 00000000001101E4
            parent_handle = ctypes.windll.user32.FindWindowW(0, "出力先のファイル名を選択してください")
            # 出力先のファイル名を選択してくださいをアクティブにする
            ctypes.windll.user32.SetForegroundWindow(parent_handle)
            
            pgui.press("tab",6,interval=0.3)
            pgui.press("enter")
            clip.copy(SetActualTimepath)
            pgui.hotkey("ctrl","v")
            # pgui.typewrite(SetActualTimepath)
            time.sleep(0.3)
            pgui.press("enter")
            time.sleep(0.3)
            pgui.hotkey("alt","s")
            print ("処理が完了しました。保存先は下記を参照。" )
            print (SetActualTimepath)
            print("-----------------------------------------------------------------------------------------")
    window.close()

    # SGvalueDat呼び出し方法
        # print(SGvalueDat["4537"]["本体シリアル番号"])
        # print(SGvalueDat["4537"]["受注番号"])
        # print(SGvalueDat["4537"]["ｴﾝﾄﾞﾕｰｻﾞｰ"])
        # print(SGvalueDat["4537"]["指図_日付"])
        # print(SGvalueDat["4537"]["指図_進捗"])
        # print(SGvalueDat["4537"]["P_日付"])
        # print(SGvalueDat["4537"]["P_進捗"])
        # print(SGvalueDat["4537"]["組_日付"])
        # print(SGvalueDat["4537"]["組_進捗"])
        # print(SGvalueDat["4537"]["立_日付"])
        # print(SGvalueDat["4537"]["立_進捗"])
        # print(SGvalueDat["4537"]["ｋ3_日付"])
        # print(SGvalueDat["4537"]["ｋ3_進捗"])
        # print(SGvalueDat["4537"]["高前_日付"])
        # print(SGvalueDat["4537"]["高前_進捗"])
        # print(SGvalueDat["4537"]["低_日付"])
        # print(SGvalueDat["4537"]["低_進捗"])
        # print(SGvalueDat["4537"]["高温_日付"])
        # print(SGvalueDat["4537"]["高温_進捗"])
        # print(SGvalueDat["4537"]["ｋ4_日付"])
        # print(SGvalueDat["4537"]["ｋ4_進捗"])
        # print(SGvalueDat["4537"]["出準_日付"])
        # print(SGvalueDat["4537"]["出準_進捗"])
        # print(SGvalueDat["4537"]["発送_日付"])
        # print(SGvalueDat["4537"]["発送_進捗"])